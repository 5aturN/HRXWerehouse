"""excel_exporter: генерация XLSX с листами «Остатки» и «Движения»."""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

HEADER_STYLE = dict(font=Font(bold=True, color="FFFFFF"),
                    fill=PatternFill("solid", fgColor="4472C4"))

async def export_report(db: AsyncSession, date_from: str, date_to: str) -> bytes:
    """Формирует Excel-файл: лист 1 — остатки, лист 2 — движения за период."""
    wb = Workbook()

    # ===== Лист «Остатки» =====
    ws = wb.active
    ws.title = "Остатки"
    headers = ["Артикул", "Наименование", "Дец. номер", "Папка", "Тип учета", "Остаток", "Порог"]
    ws.append(headers)
    rows = await db.execute(text("""
        SELECT i.sku, i.name, i.decimal_number, f.name, i.unit_type,
               COALESCE(SUM(b.remaining), 0) AS balance, i.threshold
        FROM items i
        JOIN folders f ON f.id = i.folder_id
        LEFT JOIN batches b ON b.item_id = i.id
        GROUP BY i.id ORDER BY f.name, i.name
    """))
    for r in rows:
        ws.append(list(r))
        if r[5] <= r[6]:  # подсветка критических остатков
            for cell in ws[ws.max_row]:
                cell.fill = PatternFill("solid", fgColor="FFC7CE")

    # ===== Лист «Движения за период» =====
    ws2 = wb.create_sheet("Движения за период")
    ws2.append(["Дата", "Тип операции", "Артикул", "Наименование", "Кол-во",
                "Причина/Проект", "Пользователь"])
    tx_names = {"receipt": "Приход", "assembly": "Сборка", "write_off": "Списание",
                "return": "Возврат", "scrap": "Брак", "rd_issue": "Выдача на проект"}
    rows = await db.execute(text("""
        SELECT t.created_at, t.transaction_type, i.sku, i.name, t.quantity,
               COALESCE(t.reason, pf.name, ''), COALESCE(u.full_name, t.user_id)
        FROM transactions t
        JOIN batches b ON b.id = t.batch_id
        JOIN items i ON i.id = b.item_id
        LEFT JOIN folders pf ON pf.id = t.project_id
        LEFT JOIN users u ON u.telegram_id = t.user_id
        WHERE t.created_at BETWEEN :df AND :dt
        ORDER BY t.created_at DESC
    """), {"df": date_from, "dt": date_to + " 23:59:59"})
    for r in rows:
        r = list(r); r[1] = tx_names.get(r[1], r[1])
        ws2.append(r)

    # Оформление шапок и ширина колонок
    for sheet in (ws, ws2):
        for cell in sheet[1]:
            cell.font, cell.fill = HEADER_STYLE["font"], HEADER_STYLE["fill"]
        for col in sheet.columns:
            sheet.column_dimensions[col[0].column_letter].width = \
                max(12, min(40, max(len(str(c.value or "")) for c in col) + 2))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()